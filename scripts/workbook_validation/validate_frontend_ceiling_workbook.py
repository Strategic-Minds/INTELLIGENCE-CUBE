#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import sys

try:
    from openpyxl import load_workbook
except Exception:
    print("openpyxl unavailable for standalone validation")
    sys.exit(2)

REQUIRED_SHEETS = [
    "00_START_HERE",
    "01_FRONTEND_CHARTER",
    "02_SOURCE_TRUTH_BRIDGE",
    "03_CAPABILITY_MAP",
    "04_PAGE_ARCHITECTURE",
    "05_DESIGN_TOKENS",
    "06_COMPONENT_LIBRARY",
    "07_RESPONSIVE_RULES",
    "08_ACCESS_PERF_SEO",
    "09_MOTION_STATES",
    "10_DATA_CONTRACTS",
    "11_FORMS_CONVERSION",
    "12_VERCEL_NEXT_RUNTIME",
    "13_SUPABASE_FRONTEND",
    "14_GPT_FRONTEND_DIRECTIVES",
    "15_CODEX_HANDOFF",
    "16_VALIDATION_MATRIX",
    "17_RECEIPTS_PROOF",
    "18_GAP_LEDGER",
    "19_SCORECARD",
    "20_FINAL_HANDOFF",
    "99_VALIDATION_LISTS",
]


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "01_Builder_Docs/frontend-ceiling/AUTO_BUILDER_FRONTEND_CEILING_WORKBOOK.xlsx"
    )
    wb = load_workbook(path, read_only=True, data_only=False)
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing:
        print("FRONTEND CEILING WORKBOOK VALIDATION: FAIL")
        print("missing_sheets=" + ",".join(missing))
        raise SystemExit(1)
    print("FRONTEND CEILING WORKBOOK VALIDATION: PASS")
    print(f"required_sheets={len(REQUIRED_SHEETS)}")
    print("codex_handoff=present")
    print("gpt_frontend_directives=present")


if __name__ == "__main__":
    main()

